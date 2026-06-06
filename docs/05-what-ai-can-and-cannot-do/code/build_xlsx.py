"""Generate the three Chapter 5 practitioner Excel templates.

Outputs (relative to the chapter's practitioner/ folder):
  files/ai-category-checklist.xlsx
  files/problem-first-worksheet.xlsx
  files/pre-deployment-readiness.xlsx

Each .xlsx is a working workshop tool — input cells, dropdowns,
verdict formulas, and conditional formatting. Run with the
project's Python; openpyxl is the only dependency.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

ROOT = Path(__file__).resolve().parent.parent          # practitioner/
OUT = ROOT / "files"
OUT.mkdir(exist_ok=True)

TEAL = "008B8B"
NAVY = "1F3A5F"
PALE_TEAL = "E0F2F1"
GREEN = "C8E6C9"
YELLOW = "FFF59D"
ORANGE = "FFCC80"
RED = "EF9A9A"
GREY = "ECEFF1"

THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, fill=TEAL, color="FFFFFF"):
    cell.font = Font(bold=True, color=color, size=11)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER


def style_input(cell, fill="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -----------------------------------------------------------------------------
# 1) AI Category Classification Checklist
# -----------------------------------------------------------------------------
def build_category_checklist():
    wb = Workbook()
    ws = wb.active
    ws.title = "Classifier"

    title = ws.cell(row=1, column=1, value="AI Category Classification Checklist")
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    instr = ("Answer Q1–Q3 in order. The first Yes determines the system's primary category. "
             "Answer Q4 separately — it determines whether regulated-AI overlay obligations apply.")
    ws.cell(row=2, column=1, value=instr).alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 36

    headers = ["#", "Diagnostic Question", "Answer", "If Yes → Category"]
    for c, h in enumerate(headers, 1):
        style_header(ws.cell(row=4, column=c, value=h))
    ws.row_dimensions[4].height = 26

    questions = [
        ("Q1", "Does the system execute logic that was explicitly written by a human?",
         "Rule-Based Automation"),
        ("Q2", "Does the system learn patterns from historical labeled data and produce a fixed-form output "
               "(score, classification, prediction)?", "Classical Machine Learning"),
        ("Q3", "Does the system synthesize across sources, generate language, reason through novel problems, "
               "or operate across multiple task types without retraining?", "Generative AI"),
        ("Q4", "Does the system's output directly influence a clinical decision, a regulatory submission, "
               "or a financial determination?", "Regulated AI (overlay)"),
    ]
    for i, (qid, q, cat) in enumerate(questions, start=5):
        ws.cell(row=i, column=1, value=qid).font = Font(bold=True)
        style_input(ws.cell(row=i, column=1))
        style_input(ws.cell(row=i, column=2, value=q))
        style_input(ws.cell(row=i, column=3), fill=PALE_TEAL)
        style_input(ws.cell(row=i, column=4, value=cat), fill=GREY)
        ws.row_dimensions[i].height = 44

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.add("C5:C8")
    ws.add_data_validation(dv)

    # Verdict block.
    ws.cell(row=10, column=1, value="Primary category:").font = Font(bold=True)
    verdict_formula = (
        '=IF(C5="Yes","Rule-Based Automation",'
        'IF(C6="Yes","Classical Machine Learning",'
        'IF(C7="Yes","Generative AI",'
        '"Answer Q1–Q3")))'
    )
    primary = ws.cell(row=10, column=2, value=verdict_formula)
    primary.font = Font(bold=True, size=12)
    primary.fill = PatternFill("solid", fgColor=PALE_TEAL)
    primary.border = BORDER
    ws.merge_cells("B10:D10")

    ws.cell(row=11, column=1, value="Regulated overlay:").font = Font(bold=True)
    overlay = ws.cell(
        row=11, column=2,
        value='=IF(C8="Yes","YES — add regulated-AI obligations on top of primary category","No")',
    )
    overlay.font = Font(bold=True, size=12)
    overlay.fill = PatternFill("solid", fgColor=PALE_TEAL)
    overlay.border = BORDER
    ws.merge_cells("B11:D11")

    # Color the primary verdict cell by category.
    ws.conditional_formatting.add("B10", CellIsRule(
        operator="equal", formula=['"Rule-Based Automation"'],
        fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add("B10", CellIsRule(
        operator="equal", formula=['"Classical Machine Learning"'],
        fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add("B10", CellIsRule(
        operator="equal", formula=['"Generative AI"'],
        fill=PatternFill("solid", fgColor=ORANGE)))
    ws.conditional_formatting.add("B11", CellIsRule(
        operator="beginsWith", formula=['"YES"'],
        fill=PatternFill("solid", fgColor=RED)))

    # Governance regime guidance.
    ws.cell(row=13, column=1, value="Governance regime that applies:").font = Font(bold=True, italic=True)
    ws.merge_cells("A13:D13")
    regimes = [
        ("Rule-Based Automation", "Automation governance, not AI governance."),
        ("Classical Machine Learning",
         "ML validation: training/validation/test splits, holdout-set metrics, distribution-shift monitoring."),
        ("Generative AI",
         "GenAI governance: hallucination risk management, HITL design, intended-use statement, PCCP if regulated."),
        ("Regulated AI (overlay)",
         "Layer on the applicable framework above: SaMD obligations, EU AI Act high-risk requirements, "
         "ISO 14971 risk register, intended-use documentation, audit trail."),
    ]
    for i, (cat, regime) in enumerate(regimes, start=14):
        style_input(ws.cell(row=i, column=1, value=cat), fill=GREY)
        ws.cell(row=i, column=1).font = Font(bold=True)
        style_input(ws.cell(row=i, column=2, value=regime))
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.row_dimensions[i].height = 38

    set_widths(ws, [14, 60, 26, 22])
    wb.save(OUT / "ai-category-checklist.xlsx")
    print(f"  wrote {OUT / 'ai-category-checklist.xlsx'}")


# -----------------------------------------------------------------------------
# 2) Problem-First Diagnostic Worksheet
# -----------------------------------------------------------------------------
def build_problem_first():
    wb = Workbook()
    instr = wb.active
    instr.title = "Instructions"

    title = instr.cell(row=1, column=1, value="Problem-First Diagnostic Worksheet")
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=NAVY)
    instr.merge_cells("A1:C1")
    instr.row_dimensions[1].height = 28

    steps_intro = [
        "How to use this worksheet",
        "• Run as a 90-minute session with operational leaders. No technology vendors in the room.",
        "• Complete steps in order. Steps 1–3 surface candidate problems; Step 4 prevents fragmenting "
        "  the portfolio into point solutions; Step 5 sequences the work.",
        "• Re-run annually, or whenever a new foundation-model capability tier becomes available.",
        "• If a vendor demo preceded this session, answers will reverse-engineer a problem to fit the demo. "
        "  Stop and restart from Step 1 without vendor contact.",
    ]
    for i, line in enumerate(steps_intro, start=3):
        cell = instr.cell(row=i, column=1, value=line)
        if i == 3:
            cell.font = Font(bold=True, size=12, color=TEAL)
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        instr.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        instr.row_dimensions[i].height = 22 if i == 3 else 30
    set_widths(instr, [100, 20, 20])

    # Step 1 — Effort inventory
    s1 = wb.create_sheet("1. Effort inventory")
    headers = ["Workflow", "Nature of effort (judgment / synthesis / generation / physical / relationship)",
               "Volume (decisions / day)", "Cost frame (FTE, hours, dollars)", "Notes"]
    for c, h in enumerate(headers, 1):
        style_header(s1.cell(row=1, column=c, value=h))
    s1.row_dimensions[1].height = 42
    for r in range(2, 12):
        for c in range(1, 6):
            style_input(s1.cell(row=r, column=c), fill="FFFFFF" if r % 2 == 0 else PALE_TEAL)
            s1.row_dimensions[r].height = 28
    dv = DataValidation(type="list",
                        formula1='"judgment,synthesis,generation,physical action,relationship"',
                        allow_blank=True)
    dv.add("B2:B11")
    s1.add_data_validation(dv)
    set_widths(s1, [32, 38, 18, 22, 30])

    # Step 2 — Synthesis bottlenecks
    s2 = wb.create_sheet("2. Synthesis bottlenecks")
    headers = ["Workflow (from Step 1)", "Information sources humans currently consult",
               "Sources structured? (Yes / Partial / No)",
               "Would simultaneous synthesis beat human under time pressure?"]
    for c, h in enumerate(headers, 1):
        style_header(s2.cell(row=1, column=c, value=h))
    s2.row_dimensions[1].height = 50
    for r in range(2, 12):
        for c in range(1, 5):
            style_input(s2.cell(row=r, column=c), fill="FFFFFF" if r % 2 == 0 else PALE_TEAL)
            s2.row_dimensions[r].height = 28
    dv = DataValidation(type="list", formula1='"Yes,Partial,No"', allow_blank=True)
    dv.add("C2:C11")
    s2.add_data_validation(dv)
    dv2 = DataValidation(type="list", formula1='"Yes,No,Unsure"', allow_blank=True)
    dv2.add("D2:D11")
    s2.add_data_validation(dv2)
    set_widths(s2, [32, 40, 22, 36])

    # Step 3 — Underused data
    s3 = wb.create_sheet("3. Underused data")
    headers = ["Data asset", "Operational decisions it should inform",
               "Gap driver (bandwidth / access / genuine uncertainty)",
               "Estimated decisions affected per month"]
    for c, h in enumerate(headers, 1):
        style_header(s3.cell(row=1, column=c, value=h))
    s3.row_dimensions[1].height = 44
    for r in range(2, 8):
        for c in range(1, 5):
            style_input(s3.cell(row=r, column=c), fill="FFFFFF" if r % 2 == 0 else PALE_TEAL)
            s3.row_dimensions[r].height = 30
    dv = DataValidation(type="list",
                        formula1='"bandwidth,access,genuine uncertainty"', allow_blank=True)
    dv.add("C2:C7")
    s3.add_data_validation(dv)
    set_widths(s3, [30, 38, 30, 24])

    # Step 4 — Adjacent-workflow scan
    s4 = wb.create_sheet("4. Adjacent scan")
    headers = ["Current AI investment", "Original problem solved",
               "Adjacent problems same platform could address",
               "Point-solution cost (USD)", "Platform marginal cost (USD)",
               "Cost ratio (point ÷ platform)"]
    for c, h in enumerate(headers, 1):
        style_header(s4.cell(row=1, column=c, value=h))
    s4.row_dimensions[1].height = 48
    for r in range(2, 10):
        for c in range(1, 7):
            style_input(s4.cell(row=r, column=c), fill="FFFFFF" if r % 2 == 0 else PALE_TEAL)
            s4.row_dimensions[r].height = 30
        # cost ratio formula
        s4.cell(row=r, column=6, value=f"=IFERROR(D{r}/E{r},\"\")")
    set_widths(s4, [28, 30, 36, 18, 22, 20])

    # Step 5 — Roadmap rank
    s5 = wb.create_sheet("5. Roadmap rank")
    headers = ["Opportunity", "Clinical impact (1–5)", "Financial impact (1–5)",
               "Feasibility (1–5)", "Score (C × F × X)", "Sequenced rank"]
    for c, h in enumerate(headers, 1):
        style_header(s5.cell(row=1, column=c, value=h))
    s5.row_dimensions[1].height = 38
    for r in range(2, 12):
        for c in range(1, 7):
            style_input(s5.cell(row=r, column=c), fill="FFFFFF" if r % 2 == 0 else PALE_TEAL)
            s5.row_dimensions[r].height = 28
        s5.cell(row=r, column=5, value=f"=IFERROR(B{r}*C{r}*D{r},\"\")")
        s5.cell(row=r, column=6,
                value=f'=IF(E{r}="","",RANK(E{r},$E$2:$E$11,0))')
    # Conditional formatting on score: greener = higher
    s5.conditional_formatting.add("E2:E11", CellIsRule(
        operator="greaterThanOrEqual", formula=["64"],
        fill=PatternFill("solid", fgColor=GREEN)))
    s5.conditional_formatting.add("E2:E11", CellIsRule(
        operator="between", formula=["27", "63"],
        fill=PatternFill("solid", fgColor=YELLOW)))
    s5.conditional_formatting.add("E2:E11", CellIsRule(
        operator="between", formula=["1", "26"],
        fill=PatternFill("solid", fgColor=ORANGE)))
    set_widths(s5, [36, 16, 16, 14, 18, 16])

    wb.save(OUT / "problem-first-worksheet.xlsx")
    print(f"  wrote {OUT / 'problem-first-worksheet.xlsx'}")


# -----------------------------------------------------------------------------
# 3) Pre-Deployment Readiness Assessment
# -----------------------------------------------------------------------------
def build_pre_deployment():
    wb = Workbook()
    ws = wb.active
    ws.title = "Readiness"

    title = ws.cell(row=1, column=1, value="Pre-Deployment Readiness Assessment")
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 28

    intro = ("Walk every domain before go-live. Every required item must be Confirmed, "
             "not Planned. Items left at Planned become production debt; the cost of clearing "
             "them post-launch is always higher than pre-launch.")
    ws.cell(row=2, column=1, value=intro).alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 40

    headers = ["Domain", "Required item", "Regulatory?", "Status", "Notes / evidence"]
    for c, h in enumerate(headers, 1):
        style_header(ws.cell(row=4, column=c, value=h))
    ws.row_dimensions[4].height = 26

    items = [
        ("Data Foundation",
         "Training data quality assessed and documented; missingness characterized; demographic representation audited; vocabulary version tagged",
         True),
        ("Governance Infrastructure",
         "Consent framework in place for all data sources; audit trail operational; data lineage active; RBAC implemented and tested",
         False),
        ("Hallucination Risk",
         "Hallucination failure modes identified; RAG or grounding architecture implemented; human review gates defined for all clinically significant outputs",
         True),
        ("Human Oversight Design",
         "HITL checkpoints defined and tested; override mechanism available and documented; accountability assignment documented per AI-influenced decision",
         True),
        ("Intended Use Statement",
         "Intended use documented and scoped to advisory function; decision-replacement language explicitly excluded; regulatory classification signed off",
         True),
        ("Performance Monitoring",
         "Baseline performance metrics established; drift detection in place; review cadence defined; retraining trigger criteria documented",
         False),
        ("PCCP (if applicable)",
         "Predetermined Change Control Plan drafted if continuously learning or foundation-model-based; permissible changes, performance bounds, and validation methodology specified",
         True),
    ]
    start = 5
    for i, (dom, item, reg) in enumerate(items, start=start):
        style_input(ws.cell(row=i, column=1, value=dom), fill=GREY)
        ws.cell(row=i, column=1).font = Font(bold=True)
        style_input(ws.cell(row=i, column=2, value=item))
        cell = ws.cell(row=i, column=3, value="REGULATORY" if reg else "—")
        style_input(cell, fill=RED if reg else "FFFFFF")
        if reg:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        style_input(ws.cell(row=i, column=4), fill=PALE_TEAL)
        style_input(ws.cell(row=i, column=5))
        ws.row_dimensions[i].height = 60

    end = start + len(items) - 1
    dv = DataValidation(type="list",
                        formula1='"Planned,In progress,Confirmed,Not applicable"',
                        allow_blank=True)
    dv.add(f"D{start}:D{end}")
    ws.add_data_validation(dv)

    # Conditional formatting: green=Confirmed, yellow=In progress, red=Planned
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(
        operator="equal", formula=['"Confirmed"'],
        fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(
        operator="equal", formula=['"In progress"'],
        fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(
        operator="equal", formula=['"Planned"'],
        fill=PatternFill("solid", fgColor=RED)))

    # Go/No-Go: GO only if every REGULATORY item is Confirmed.
    verdict_row = end + 2
    ws.cell(row=verdict_row, column=1, value="Go / No-Go:").font = Font(bold=True, size=12)
    verdict_formula = (
        '=IF(COUNTIFS(C{s}:C{e},"REGULATORY",D{s}:D{e},"Confirmed")='
        'COUNTIF(C{s}:C{e},"REGULATORY"),"GO","NO-GO — regulatory items still open")'
    ).format(s=start, e=end)
    v = ws.cell(row=verdict_row, column=2, value=verdict_formula)
    v.font = Font(bold=True, size=14)
    v.fill = PatternFill("solid", fgColor=PALE_TEAL)
    v.border = BORDER
    v.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=verdict_row, start_column=2, end_row=verdict_row, end_column=5)
    ws.row_dimensions[verdict_row].height = 32

    ws.conditional_formatting.add(
        f"B{verdict_row}",
        FormulaRule(formula=[f'B{verdict_row}="GO"'],
                    fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        f"B{verdict_row}",
        FormulaRule(formula=[f'LEFT(B{verdict_row},2)="NO"'],
                    fill=PatternFill("solid", fgColor=RED)),
    )

    set_widths(ws, [22, 56, 14, 16, 32])
    wb.save(OUT / "pre-deployment-readiness.xlsx")
    print(f"  wrote {OUT / 'pre-deployment-readiness.xlsx'}")


if __name__ == "__main__":
    print(f"Generating Chapter 5 practitioner Excel templates -> {OUT}")
    build_category_checklist()
    build_problem_first()
    build_pre_deployment()
    print("Done.")
